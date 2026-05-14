import sys
import time
import random
import logging
import io
from datetime import datetime
from pathlib import Path

import streamlit as st
import pandas as pd

sys.path.insert(0, str(Path(__file__).parent))

from config.settings import (
    OUTPUT_DIR,
    LOG_DIR,
    MAX_LOCATIONS,
    REQUEST_DELAY_MIN,
    REQUEST_DELAY_MAX,
    LONG_DELAY_MIN,
    LONG_DELAY_MAX,
    LONG_DELAY_EVERY_N,
)
from scraper.discovery import discover_locator_url
from scraper.renderer import render_page
from scraper.extractor import extract_locations, probe_direct_apis
from scraper.cleaner import clean_locations
from scraper.exporter import export_results, COLUMNS


# ── Page config ───────────────────────────────────────────────────────

st.set_page_config(
    page_title="Location Scraper",
    page_icon="logo.svg",
    layout="wide",
    initial_sidebar_state="expanded",
)


# ── Custom CSS ────────────────────────────────────────────────────────

st.markdown("""
<style>
    /* INDATEL Labs palette
       --indatel-navy:    #012169  (Pantone 280, primary)
       --indatel-navy-2:  #001A52  (deeper navy)
       --indatel-blue:    #0A4FB8  (accent blue)
       --indatel-gray:    #898D8D  (Pantone 423)
       --indatel-light:   #F4F6FA  (page background tint)
    */

    /* Main container */
    .main .block-container {
        padding-top: 2rem;
        max-width: 1200px;
    }

    /* Header */
    .app-header {
        background: linear-gradient(135deg, #001A52 0%, #012169 55%, #0A4FB8 100%);
        padding: 2rem 2.5rem;
        border-radius: 12px;
        margin-bottom: 2rem;
        box-shadow: 0 6px 24px rgba(1, 33, 105, 0.20);
        position: relative;
        overflow: hidden;
    }
    .app-header::after {
        content: "";
        position: absolute;
        inset: 0;
        background: radial-gradient(circle at 90% 20%, rgba(255,255,255,0.08), transparent 60%);
        pointer-events: none;
    }
    .app-header h1 {
        color: white;
        font-size: 2rem;
        font-weight: 700;
        margin: 0 0 0.3rem 0;
        letter-spacing: -0.02em;
    }
    .app-header p {
        color: rgba(255,255,255,0.85);
        font-size: 1rem;
        margin: 0;
    }
    .app-header .brand-tag {
        display: inline-block;
        margin-top: 0.75rem;
        padding: 0.2rem 0.7rem;
        background: rgba(255,255,255,0.12);
        border: 1px solid rgba(255,255,255,0.25);
        border-radius: 999px;
        color: #FFFFFF;
        font-size: 0.75rem;
        font-weight: 600;
        letter-spacing: 0.08em;
        text-transform: uppercase;
    }

    /* Metric cards */
    .metric-row {
        display: flex;
        gap: 1rem;
        margin: 1.5rem 0;
    }
    .metric-card {
        background: #FFFFFF;
        border: 1px solid #E2E6EE;
        border-radius: 10px;
        padding: 1.25rem 1.5rem;
        flex: 1;
        text-align: center;
        transition: border-color 0.2s, box-shadow 0.2s;
    }
    .metric-card:hover {
        border-color: #012169;
        box-shadow: 0 4px 12px rgba(1, 33, 105, 0.10);
    }
    .metric-card .metric-value {
        font-size: 2rem;
        font-weight: 700;
        color: #001A52;
        line-height: 1.2;
    }
    .metric-card .metric-label {
        font-size: 0.8rem;
        color: #6B7280;
        text-transform: uppercase;
        letter-spacing: 0.05em;
        margin-top: 0.25rem;
    }
    .metric-card.accent .metric-value { color: #012169; }
    .metric-card.success .metric-value { color: #15803D; }
    .metric-card.warning .metric-value { color: #B45309; }
    .metric-card.error .metric-value { color: #B91C1C; }

    /* Status log */
    .status-log {
        background: #F4F6FA;
        border: 1px solid #E2E6EE;
        border-radius: 8px;
        padding: 1rem;
        font-family: 'Consolas', 'Monaco', monospace;
        font-size: 0.8rem;
        max-height: 300px;
        overflow-y: auto;
        color: #3F4754;
        line-height: 1.6;
    }
    .status-log .log-info { color: #012169; }
    .status-log .log-success { color: #15803D; }
    .status-log .log-warning { color: #B45309; }
    .status-log .log-error { color: #B91C1C; }
    .status-log .log-step { color: #0A4FB8; font-weight: 600; }

    /* Sidebar */
    section[data-testid="stSidebar"] {
        background: #F4F6FA;
        border-right: 1px solid #E2E6EE;
    }
    section[data-testid="stSidebar"] h3 {
        color: #012169;
    }

    /* Primary button */
    .stButton > button {
        background: linear-gradient(135deg, #012169, #0A4FB8);
        color: white;
        border: none;
        padding: 0.6rem 2rem;
        font-weight: 600;
        border-radius: 8px;
        transition: all 0.2s;
        width: 100%;
    }
    .stButton > button:hover {
        box-shadow: 0 4px 16px rgba(1, 33, 105, 0.35);
        transform: translateY(-1px);
    }

    /* Download button */
    .stDownloadButton > button {
        background: #FFFFFF;
        border: 1px solid #E2E6EE;
        color: #001A52;
        border-radius: 8px;
        transition: all 0.2s;
    }
    .stDownloadButton > button:hover {
        border-color: #012169;
        background: #EFF3FB;
    }

    /* Dataframe */
    .stDataFrame {
        border-radius: 10px;
        overflow: hidden;
        border: 1px solid #E2E6EE;
    }

    /* Tabs */
    .stTabs [data-baseweb="tab-list"] {
        gap: 0.5rem;
    }
    .stTabs [data-baseweb="tab"] {
        border-radius: 8px 8px 0 0;
        padding: 0.5rem 1.5rem;
    }
    .stTabs [aria-selected="true"] {
        color: #012169 !important;
    }

    /* Inputs */
    .stTextInput > div > div > input {
        background: #FFFFFF;
        border-color: #E2E6EE;
        color: #001A52;
        border-radius: 8px;
    }
    .stTextInput > div > div > input:focus {
        border-color: #012169;
        box-shadow: 0 0 0 1px #012169;
    }

    /* Expander */
    .streamlit-expanderHeader {
        background: #F4F6FA;
        border-radius: 8px;
    }

    /* Section headers */
    .main .stMarkdown h3 {
        color: #001A52;
        font-weight: 600;
        border-bottom: 2px solid #E2E6EE;
        padding-bottom: 0.5rem;
        margin-top: 1.5rem;
    }

    /* Footer brand line */
    .footer-brand {
        margin-top: 3rem;
        padding-top: 1.25rem;
        border-top: 1px solid #E2E6EE;
        text-align: center;
        color: #898D8D;
        font-size: 0.78rem;
        letter-spacing: 0.05em;
    }
    .footer-brand strong {
        color: #012169;
        font-weight: 700;
        letter-spacing: 0.08em;
    }

    /* Hide streamlit branding */
    #MainMenu { visibility: hidden; }
    footer { visibility: hidden; }
    header { visibility: hidden; }
    /* Keep the sidebar-expand arrow visible when sidebar is collapsed —
       it lives inside the (otherwise hidden) header. */
    header [data-testid="stExpandSidebarButton"],
    button[data-testid="stExpandSidebarButton"] {
        visibility: visible !important;
        z-index: 999999;
    }
</style>
""", unsafe_allow_html=True)


# ── Session state init ────────────────────────────────────────────────

if "results" not in st.session_state:
    st.session_state.results = None
if "log_lines" not in st.session_state:
    st.session_state.log_lines = []
if "is_running" not in st.session_state:
    st.session_state.is_running = False
if "run_stats" not in st.session_state:
    st.session_state.run_stats = {}


# ── Logging bridge ────────────────────────────────────────────────────

class StreamlitLogHandler(logging.Handler):
    def __init__(self, log_container):
        super().__init__()
        self.log_container = log_container

    def emit(self, record):
        msg = self.format(record)
        level = record.levelname
        if level == "INFO":
            css = "log-info"
        elif level == "WARNING":
            css = "log-warning"
        elif level == "ERROR":
            css = "log-error"
        else:
            css = ""

        if "Step" in msg and ":" in msg:
            css = "log-step"

        st.session_state.log_lines.append(f'<span class="{css}">{msg}</span>')
        # Keep last 200 lines
        if len(st.session_state.log_lines) > 200:
            st.session_state.log_lines = st.session_state.log_lines[-200:]


def setup_ui_logging():
    fmt = logging.Formatter("[%(asctime)s] %(message)s", datefmt="%H:%M:%S")
    handler = StreamlitLogHandler(None)
    handler.setFormatter(fmt)
    handler.setLevel(logging.INFO)

    root = logging.getLogger()
    root.setLevel(logging.INFO)
    # Remove existing handlers to avoid duplicates
    for h in root.handlers[:]:
        root.removeHandler(h)
    root.addHandler(handler)

    # Also log to file
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    log_file = LOG_DIR / f"scraper_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    file_handler = logging.FileHandler(log_file, encoding="utf-8")
    file_handler.setFormatter(logging.Formatter(
        "[%(asctime)s] [%(levelname)s] [%(name)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    ))
    file_handler.setLevel(logging.DEBUG)
    root.addHandler(file_handler)


# ── Scraping logic ────────────────────────────────────────────────────

def scrape_company(company_name: str, progress_callback=None, manual_url: str | None = None) -> list[dict]:
    logger = logging.getLogger("scraper")

    if manual_url:
        url, method = manual_url.strip(), "manual"
        logger.info("[%s] Using manual URL: %s", company_name, url)
    else:
        if progress_callback:
            progress_callback("discovery", company_name)
        url, method = discover_locator_url(company_name)
        if not url:
            logger.warning("[%s] Could not find store locator URL", company_name)
            return []
        logger.info("[%s] Found URL via %s: %s", company_name, method, url)

    time.sleep(random.uniform(REQUEST_DELAY_MIN, REQUEST_DELAY_MAX))

    if progress_callback:
        progress_callback("rendering", company_name)

    result = render_page(url)
    if not result:
        logger.warning("[%s] Failed to render page", company_name)
        return []
    logger.info("[%s] Rendered via %s (%d bytes)", company_name, result.method, len(result.html))

    if progress_callback:
        progress_callback("extracting", company_name)

    # Forward the directory-crawl's phase progress to the UI callback.
    def _dir_progress(phase, current, total):
        if not progress_callback:
            return
        if phase == "estimate":
            # Pass the count as a plain integer string the UI can parse.
            progress_callback("crawl_estimate", company_name, str(total))
            return
        if total:
            pct = (current / total) * 100
            progress_callback(f"crawl_{phase}", company_name, f"{current:,}/{total:,} ({pct:.0f}%)")

    raw_locations = extract_locations(result, company_name, progress_cb=_dir_progress)

    if not raw_locations and result.method == "static":
        logger.info("[%s] Retrying with Playwright...", company_name)
        time.sleep(random.uniform(REQUEST_DELAY_MIN, REQUEST_DELAY_MAX))
        pw_result = render_page(url, force_playwright=True)
        if pw_result:
            raw_locations = extract_locations(pw_result, company_name, progress_cb=_dir_progress)

    # Try the locations.{brand}.com subdomain (Yext/Uberall directory pattern)
    if not raw_locations:
        from scraper.discovery import _slug_candidates
        from scraper.extractor import extract_from_directory_tree
        import requests
        for slug in _slug_candidates(company_name):
            alt_url = f"https://locations.{slug}.com/index.html"
            try:
                logger.info("[%s] Trying directory subdomain: %s", company_name, alt_url)
                r = requests.get(alt_url, headers={"User-Agent": "Mozilla/5.0"}, timeout=12, allow_redirects=True)
                if r.status_code != 200 or len(r.text) < 1000:
                    continue
                dir_locs = extract_from_directory_tree(r.text, alt_url, progress_cb=_dir_progress)
                if dir_locs:
                    raw_locations = dir_locs
                    logger.info("[%s] Directory subdomain: %d locations", company_name, len(dir_locs))
                    break
            except requests.RequestException:
                continue

    if not raw_locations:
        logger.info("[%s] Trying direct API probing...", company_name)
        raw_locations = probe_direct_apis(url, company_name)

    if not raw_locations:
        logger.warning("[%s] No locations extracted", company_name)
        return []

    logger.info("[%s] Extracted %d raw locations", company_name, len(raw_locations))

    if progress_callback:
        progress_callback("cleaning", company_name)

    cleaned = clean_locations(raw_locations, company_name)
    logger.info("[%s] Final: %d locations", company_name, len(cleaned))

    if MAX_LOCATIONS > 0 and len(cleaned) > MAX_LOCATIONS:
        cleaned = cleaned[:MAX_LOCATIONS]

    return cleaned


def run_scraper(companies: list[str], progress_bar, status_text, step_container, manual_url: str | None = None):
    setup_ui_logging()
    logger = logging.getLogger("main")
    st.session_state.log_lines = []

    all_locations = []
    errors = []
    start_time = time.time()

    # Per-company state captured by the progress callback so we can show
    # a friendlier estimate-and-ETA banner before the long phase 3 starts.
    estimate_state: dict[str, int] = {}

    def progress_callback(step, company, detail: str = ""):
        step_labels = {
            "discovery": "Discovering store locator URL",
            "rendering": "Rendering page",
            "extracting": "Extracting locations",
            "cleaning": "Cleaning & normalizing",
            "crawl_states": "Walking states",
            "crawl_cities": "Walking cities",
            "crawl_stores": "Fetching store pages",
        }
        # Catch the "estimate" event from the directory crawl
        if step == "crawl_estimate":
            try:
                total = int(detail.replace(",", ""))
            except Exception:
                total = 0
            estimate_state[company] = total
            # Calibrated rate: ~10 pages/sec across cities + stores combined.
            # Add 25% buffer for safety.
            est_seconds = (total / 10.0) * 1.25
            mins = int(est_seconds // 60)
            secs = int(est_seconds % 60)
            eta = f"~{mins}m {secs:02d}s" if mins else f"~{secs}s"
            status_text.markdown(
                f"**{company}** — Found `{total:,}` expected locations · "
                f"estimated runtime **{eta}** (with 25% buffer)"
            )
            return

        label = step_labels.get(step, step)
        if detail:
            status_text.markdown(f"**{company}** — {label} · `{detail}`")
        else:
            status_text.markdown(f"**{company}** — {label}")

    for i, company in enumerate(companies):
        progress = (i) / len(companies)
        progress_bar.progress(progress, text=f"Processing {i+1}/{len(companies)}: {company}")

        try:
            locations = scrape_company(company, progress_callback, manual_url=manual_url)
            all_locations.extend(locations)
            if not locations:
                errors.append((company, "No locations found"))
        except Exception as e:
            logger.error("[%s] Error: %s", company, e)
            errors.append((company, str(e)))

        if i < len(companies) - 1:
            if (i + 1) % LONG_DELAY_EVERY_N == 0:
                delay = random.uniform(LONG_DELAY_MIN, LONG_DELAY_MAX)
            else:
                delay = random.uniform(REQUEST_DELAY_MIN, REQUEST_DELAY_MAX)
            time.sleep(delay)

    progress_bar.progress(1.0, text="Complete")
    elapsed = time.time() - start_time

    # Save results
    st.session_state.run_stats = {
        "total_companies": len(companies),
        "successful": len(companies) - len(errors),
        "total_locations": len(all_locations),
        "errors": errors,
        "elapsed": elapsed,
    }

    if all_locations:
        df = pd.DataFrame(all_locations)
        for col in COLUMNS:
            if col not in df.columns:
                df[col] = ""
        df = df[COLUMNS]
        st.session_state.results = df

        base_name = f"locations_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        export_results(all_locations, "both", base_name)
    else:
        st.session_state.results = pd.DataFrame()

    return all_locations


# ── Header ────────────────────────────────────────────────────────────

st.markdown("""
<div class="app-header">
    <span class="brand-tag">By INDATEL Labs</span>
    <h1 style="margin-top: 0.6rem;">Location Scraper</h1>
    <p>Extract physical locations for any retail chain from their official website</p>
</div>
""", unsafe_allow_html=True)


# ── Sidebar ───────────────────────────────────────────────────────────

with st.sidebar:
    st.markdown("### Configuration")

    output_format = st.selectbox(
        "Output Format",
        ["Both (CSV + Excel)", "CSV Only", "Excel Only"],
        index=0,
    )

    st.markdown("---")
    st.markdown("### Tips")
    st.markdown("""
    - Use the **exact company name** as it appears on their website
    - Works best with US-based retail chains
    - Search-based locators may return partial results
    - Large chains (1000+ locations) may take 1-2 minutes
    """)


# ── Main content ──────────────────────────────────────────────────────

tab_single, tab_batch = st.tabs(["Single Company", "Batch (Excel Upload)"])

with tab_single:
    # Wrapping in a form so pressing Enter in the text field submits.
    with st.form(key="single_form", clear_on_submit=False, border=False):
        col_input, col_btn = st.columns([3, 1])
        with col_input:
            company_name = st.text_input(
                "Company Name",
                placeholder="e.g. Aldi, Starbucks, Dollar Tree...",
                label_visibility="collapsed",
                key="single_company_input",
            )
        with col_btn:
            st.markdown("<div style='height: 0.45rem'></div>", unsafe_allow_html=True)
            run_single = st.form_submit_button(
                "Scrape Locations", use_container_width=True
            )
        manual_url = st.text_input(
            "Locator URL (optional)",
            placeholder="https://locations.example.com/  — paste if auto-discovery picks the wrong page",
            key="single_manual_url",
            help="If you already know the chain's store locator URL, paste it here to skip auto-discovery. Useful for chains where the scraper finds the wrong page.",
        )

with tab_batch:
    uploaded_file = st.file_uploader(
        "Upload Excel file with company names in the first column",
        type=["xlsx"],
        key="batch_upload",
    )

    if uploaded_file:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(uploaded_file.read()), read_only=True)
        ws = wb.active
        companies_from_file = []
        for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
            val = row[0]
            if val and isinstance(val, str):
                cleaned = val.strip()
                if cleaned and cleaned.lower() not in ("company", "company_name", "name", "company name"):
                    companies_from_file.append(cleaned)
        wb.close()

        if companies_from_file:
            st.success(f"Found **{len(companies_from_file)}** companies: {', '.join(companies_from_file[:10])}")
            if len(companies_from_file) > 10:
                st.caption(f"...and {len(companies_from_file) - 10} more")
        else:
            st.error("No company names found in the first column.")
            companies_from_file = []

    run_batch = st.button("Scrape All Companies", key="run_batch", use_container_width=True)


# ── Execute scraping ──────────────────────────────────────────────────

companies_to_scrape = []
single_manual_url: str | None = None


def _company_from_url(u: str) -> str:
    """Derive a friendly company name from a locator URL.
    locations.dollartree.com → "Dollar Tree"
    www.starbucks.com/store-locator → "Starbucks"
    """
    from urllib.parse import urlparse
    host = (urlparse(u).hostname or "").lower()
    if host.startswith("www."):
        host = host[4:]
    if host.startswith(("locations.", "stores.", "find.", "store.", "restaurants.")):
        host = host.split(".", 1)[1]
    parts = host.split(".")
    brand = parts[-2] if len(parts) >= 2 else (parts[0] if parts else "company")
    # CamelCase: dollartree → Dollar Tree, sevenseleven → Sevenseleven (best effort)
    return brand.replace("-", " ").title() if brand else "Company"


if run_single:
    url_in = (manual_url or "").strip()
    name_in = (company_name or "").strip()
    if name_in:
        companies_to_scrape = [name_in]
        single_manual_url = url_in or None
    elif url_in:
        # URL provided but no name — derive a name and use the URL directly
        companies_to_scrape = [_company_from_url(url_in)]
        single_manual_url = url_in
elif run_batch and uploaded_file and companies_from_file:
    companies_to_scrape = companies_from_file

if companies_to_scrape:
    st.session_state.is_running = True
    st.session_state.results = None
    st.session_state.run_stats = {}

    st.markdown("---")
    st.markdown("### Scraping Progress")

    progress_bar = st.progress(0, text="Starting...")
    status_text = st.empty()
    step_container = st.empty()

    with st.spinner(""):
        run_scraper(companies_to_scrape, progress_bar, status_text, step_container, manual_url=single_manual_url)

    st.session_state.is_running = False
    status_text.empty()


# ── Results display ───────────────────────────────────────────────────

if st.session_state.run_stats:
    stats = st.session_state.run_stats
    st.markdown("---")

    # Metric cards
    err_count = len(stats.get("errors", []))
    st.markdown(f"""
    <div class="metric-row">
        <div class="metric-card accent">
            <div class="metric-value">{stats.get('successful', 0)}/{stats.get('total_companies', 0)}</div>
            <div class="metric-label">Companies Processed</div>
        </div>
        <div class="metric-card success">
            <div class="metric-value">{stats.get('total_locations', 0):,}</div>
            <div class="metric-label">Locations Found</div>
        </div>
        <div class="metric-card {'warning' if err_count else 'success'}">
            <div class="metric-value">{err_count}</div>
            <div class="metric-label">Errors</div>
        </div>
        <div class="metric-card">
            <div class="metric-value">{stats.get('elapsed', 0):.0f}s</div>
            <div class="metric-label">Time Elapsed</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    if stats.get("errors"):
        with st.expander(f"Errors ({len(stats['errors'])})"):
            for comp, err in stats["errors"]:
                st.markdown(f"- **{comp}**: {err}")


if st.session_state.results is not None and not st.session_state.results.empty:
    df = st.session_state.results

    st.markdown("### Results")

    # Filters
    col_f1, col_f2, col_f3, col_f4 = st.columns(4)
    with col_f1:
        companies = sorted(df["company_name"].unique())
        selected_company = st.selectbox("Company", ["All"] + companies, key="filter_company")
    with col_f2:
        states = sorted(df["state"].dropna().unique())
        selected_state = st.selectbox("State", ["All"] + [s for s in states if s], key="filter_state")
    with col_f3:
        quality_opts = sorted(df["data_quality"].dropna().unique())
        selected_quality = st.selectbox("Data Quality", ["All"] + list(quality_opts), key="filter_quality")
    with col_f4:
        search_text = st.text_input("Search", placeholder="Filter by address, city...", key="filter_search")

    filtered = df.copy()
    if selected_company != "All":
        filtered = filtered[filtered["company_name"] == selected_company]
    if selected_state != "All":
        filtered = filtered[filtered["state"] == selected_state]
    if selected_quality != "All":
        filtered = filtered[filtered["data_quality"] == selected_quality]
    if search_text:
        mask = filtered.astype(str).apply(lambda row: row.str.contains(search_text, case=False, na=False).any(), axis=1)
        filtered = filtered[mask]

    st.caption(f"Showing {len(filtered):,} of {len(df):,} locations")

    display_cols = [
        "company_name", "location_name", "street_address", "city", "state",
        "zip_code", "phone_number", "data_quality",
    ]
    st.dataframe(
        filtered[display_cols],
        use_container_width=True,
        height=450,
        column_config={
            "company_name": st.column_config.TextColumn("Company", width="medium"),
            "location_name": st.column_config.TextColumn("Location", width="medium"),
            "street_address": st.column_config.TextColumn("Address", width="large"),
            "city": st.column_config.TextColumn("City", width="small"),
            "state": st.column_config.TextColumn("State", width="small"),
            "zip_code": st.column_config.TextColumn("ZIP", width="small"),
            "phone_number": st.column_config.TextColumn("Phone", width="medium"),
            "data_quality": st.column_config.TextColumn("Quality", width="small"),
        },
    )

    # State distribution chart
    if len(filtered) > 10:
        with st.expander("Location Distribution by State"):
            state_counts = filtered["state"].value_counts().head(20)
            st.bar_chart(state_counts)

    # Download buttons
    st.markdown("### Download Results")
    col_d1, col_d2, col_d3 = st.columns(3)

    with col_d1:
        csv_data = filtered.to_csv(index=False).encode("utf-8")
        st.download_button(
            label="Download CSV",
            data=csv_data,
            file_name=f"locations_{datetime.now().strftime('%Y%m%d')}.csv",
            mime="text/csv",
            use_container_width=True,
        )

    with col_d2:
        excel_buffer = io.BytesIO()
        filtered.to_excel(excel_buffer, index=False, engine="openpyxl")
        excel_buffer.seek(0)
        st.download_button(
            label="Download Excel",
            data=excel_buffer.getvalue(),
            file_name=f"locations_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    with col_d3:
        full_csv = df.to_csv(index=False).encode("utf-8")
        st.download_button(
            label="Download All (Unfiltered)",
            data=full_csv,
            file_name=f"all_locations_{datetime.now().strftime('%Y%m%d')}.csv",
            mime="text/csv",
            use_container_width=True,
        )

    # Activity log
    if st.session_state.log_lines:
        with st.expander("Activity Log"):
            log_html = "<br>".join(st.session_state.log_lines[-100:])
            st.markdown(f'<div class="status-log">{log_html}</div>', unsafe_allow_html=True)

elif st.session_state.results is not None and st.session_state.results.empty:
    st.info("No locations were found. Try a different company name or check the activity log for details.")
    if st.session_state.log_lines:
        with st.expander("Activity Log"):
            log_html = "<br>".join(st.session_state.log_lines[-100:])
            st.markdown(f'<div class="status-log">{log_html}</div>', unsafe_allow_html=True)


# ── Footer ────────────────────────────────────────────────────────────

st.markdown("""
<div class="footer-brand">
    By <strong>INDATEL LABS</strong>
</div>
""", unsafe_allow_html=True)
