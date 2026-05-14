"""
CLI for the AI-powered Location Scraper.

  py main_ai.py --company "Sprouts"
  py main_ai.py --company "Joe's Pizza" --gemini-key YOUR_KEY
  py main_ai.py --file companies.xlsx
"""

import argparse
import logging
import os
import random
import sys
import time
from datetime import datetime
from pathlib import Path

# Reach the unmodified parent scraper package
_PARENT = Path(__file__).resolve().parent.parent
if str(_PARENT) not in sys.path:
    sys.path.insert(0, str(_PARENT))
_HERE = Path(__file__).resolve().parent
if str(_HERE) not in sys.path:
    sys.path.insert(0, str(_HERE))

from config.settings import (
    OUTPUT_DIR, LOG_DIR, LOG_LEVEL,
    REQUEST_DELAY_MIN, REQUEST_DELAY_MAX,
    LONG_DELAY_MIN, LONG_DELAY_MAX, LONG_DELAY_EVERY_N,
)
from scraper.exporter import export_results
from scraper_ai.orchestrator import scrape_company_ai
from scraper_ai.planner import Planner, PlannerConfig


def setup_logging():
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    log_file = LOG_DIR / f"scraper_ai_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    fmt = logging.Formatter(
        "[%(asctime)s] [%(levelname)s] [%(name)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    fh = logging.FileHandler(log_file, encoding="utf-8")
    fh.setFormatter(fmt); fh.setLevel(logging.DEBUG)
    sh = logging.StreamHandler(sys.stdout)
    sh.setFormatter(fmt); sh.setLevel(getattr(logging, LOG_LEVEL, logging.INFO))
    root = logging.getLogger()
    root.setLevel(logging.DEBUG)
    root.addHandler(fh); root.addHandler(sh)
    return log_file


def load_companies(filepath: str) -> list[str]:
    from openpyxl import load_workbook
    path = Path(filepath)
    if not path.exists():
        print(f"Error: File not found: {filepath}")
        sys.exit(1)
    wb = load_workbook(path, read_only=True); ws = wb.active
    out = []
    for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
        v = row[0]
        if v and isinstance(v, str):
            vc = v.strip()
            if vc and vc.lower() not in ("company", "company_name", "name", "company name"):
                out.append(vc)
    wb.close()
    if not out:
        print("Error: No company names found in the first column.")
        sys.exit(1)
    return out


def main():
    p = argparse.ArgumentParser(description="Location Scraper — AI-powered")
    p.add_argument("--company", "-c")
    p.add_argument("--file", "-f")
    p.add_argument("--url", help="Optional locator URL (skip discovery)")
    p.add_argument("--gemini-key", help="Gemini API key (or set GEMINI_API_KEY env)")
    p.add_argument("--format", choices=["csv", "excel", "both"], default="both")
    args = p.parse_args()

    if not args.company and not args.file:
        p.print_help()
        sys.exit(1)

    api_key = args.gemini_key or os.getenv("GEMINI_API_KEY") or os.getenv("GOOGLE_API_KEY")
    planner = Planner(PlannerConfig(api_key=api_key)) if api_key else None

    log_file = setup_logging()
    logger = logging.getLogger("main")
    logger.info("Location Scraper (AI) started — planner=%s", "Gemini" if planner else "OFF")
    logger.info("Log: %s", log_file)

    companies = [args.company] if args.company else load_companies(args.file)

    all_locations, errors = [], []
    start = time.time()
    for i, company in enumerate(companies):
        try:
            locs = scrape_company_ai(company, planner=planner,
                                      manual_url=args.url if i == 0 else None)
            all_locations.extend(locs)
            if not locs:
                errors.append((company, "No locations found"))
        except Exception as e:
            logger.error("[%s] %s", company, e, exc_info=True)
            errors.append((company, str(e)))
        if i < len(companies) - 1:
            delay = random.uniform(LONG_DELAY_MIN, LONG_DELAY_MAX) if (i + 1) % LONG_DELAY_EVERY_N == 0 \
                else random.uniform(REQUEST_DELAY_MIN, REQUEST_DELAY_MAX)
            time.sleep(delay)

    elapsed = time.time() - start
    if all_locations:
        base = f"locations_ai_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        out = export_results(all_locations, args.format, base)
        output_str = ", ".join(str(v) for v in out.values())
    else:
        output_str = "(no data)"

    print()
    print("=" * 50)
    print(f"  Companies processed: {len(companies) - len(errors)}/{len(companies)}")
    print(f"  Total locations: {len(all_locations):,}")
    if errors:
        print(f"  Errors ({len(errors)}):")
        for c, e in errors:
            print(f"    - {c}: {e}")
    print(f"  Output: {output_str}")
    print(f"  Elapsed: {elapsed:.1f}s")
    print("=" * 50)


if __name__ == "__main__":
    main()
