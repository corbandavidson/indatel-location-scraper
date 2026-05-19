"""
Location Scraper — AI-powered variant.

Same UX as the legacy app, but uses an AI planner to pick the
locator URL and extraction strategy. Falls back to the legacy pipeline
when no API key is configured, so the app still works out of the box.
"""

import sys
import time
import random
import logging
import io
import os
import json
import html as html_mod
from datetime import datetime
from pathlib import Path

import streamlit as st
import pandas as pd

# Local — single source of truth for the app version
_HERE_FOR_VER = Path(__file__).resolve().parent
if str(_HERE_FOR_VER) not in sys.path:
    sys.path.insert(0, str(_HERE_FOR_VER))
from version import __version__, GITHUB_REPO


@st.cache_data(ttl=3600, show_spinner=False)
def _fetch_latest_release() -> dict:
    """
    Ask GitHub for the latest release of this app. Cached for an hour so we
    don't hammer the API on every Streamlit rerun. Returns {} on any failure
    so the rest of the UI silently moves on.
    """
    import urllib.request

    try:
        req = urllib.request.Request(
            f"https://api.github.com/repos/{GITHUB_REPO}/releases/latest",
            headers={
                "Accept": "application/vnd.github+json",
                "User-Agent": "LocationScraperAI",
            },
        )
        with urllib.request.urlopen(req, timeout=4) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception:
        return {}
    tag = (data.get("tag_name") or "").lstrip("v").strip()
    if not tag:
        return {}
    return {
        "version": tag,
        "url": data.get("html_url") or f"https://github.com/{GITHUB_REPO}/releases",
    }


def _version_tuple(v: str) -> tuple:
    """Parse '1.2.3' into (1,2,3). Non-numeric segments collapse to 0 so a
    malformed tag never makes us claim an update is available."""
    out = []
    for part in v.split("."):
        try:
            out.append(int(part))
        except ValueError:
            out.append(0)
    return tuple(out)


def check_for_update() -> tuple[bool, str, str]:
    """Returns (update_available, latest_version, release_url)."""
    latest = _fetch_latest_release()
    if not latest:
        return False, "", ""
    latest_v = latest.get("version", "")
    if not latest_v:
        return False, "", ""
    return (
        _version_tuple(latest_v) > _version_tuple(__version__),
        latest_v,
        latest.get("url", ""),
    )


# ── Persistent settings ───────────────────────────────────────────────

SETTINGS_DIR = Path.home() / ".location_scraper"
SETTINGS_FILE = SETTINGS_DIR / "settings.json"
DEFAULT_SETTINGS = {
    "gemini_api_key": "",
    "output_format": "Both (CSV + Excel)",
    "gemini_model": "gemini-2.5-flash",
}


def load_settings() -> dict:
    """Load settings from ~/.location_scraper/settings.json, merging defaults."""
    out = dict(DEFAULT_SETTINGS)
    try:
        if SETTINGS_FILE.exists():
            stored = json.loads(SETTINGS_FILE.read_text(encoding="utf-8"))
            if isinstance(stored, dict):
                out.update({k: v for k, v in stored.items() if k in DEFAULT_SETTINGS})
    except Exception:
        pass
    return out


def save_settings(d: dict) -> None:
    """Persist settings. Warns in the UI on failure so users know their
    settings won't survive a restart."""
    try:
        SETTINGS_DIR.mkdir(parents=True, exist_ok=True)
        SETTINGS_FILE.write_text(json.dumps(d, indent=2), encoding="utf-8")
    except Exception:
        st.warning(
            f"Could not save settings to {SETTINGS_FILE}. "
            "Your changes will be lost when the app restarts.",
            icon="⚠️",
        )

# Make the legacy scraper modules importable
_PARENT = Path(__file__).resolve().parent.parent
if str(_PARENT) not in sys.path:
    sys.path.insert(0, str(_PARENT))
# And this folder, for scraper_ai
_HERE = Path(__file__).resolve().parent
if str(_HERE) not in sys.path:
    sys.path.insert(0, str(_HERE))

from config.settings import (
    OUTPUT_DIR,
    LOG_DIR,
    MAX_LOCATIONS,
    REQUEST_DELAY_MIN,
    REQUEST_DELAY_MAX,
    LONG_DELAY_MIN,
    LONG_DELAY_MAX,
    LONG_DELAY_EVERY_N,
)
from scraper.exporter import export_results, COLUMNS
from scraper_ai.orchestrator import scrape_company_ai
from scraper_ai.planner import Planner, PlannerConfig


# ── Page config ───────────────────────────────────────────────────────

st.set_page_config(
    page_title="Location Scraper — AI",
    page_icon="logo.svg" if (_PARENT / "logo.svg").exists() else None,
    layout="wide",
    initial_sidebar_state="expanded",
)


# ── Custom CSS (INDATEL Labs palette) ─────────────────────────────────

st.markdown("""
<style>
    .main .block-container { padding-top: 2rem; max-width: 1200px; }

    .app-header {
        background: linear-gradient(135deg, #001A52 0%, #012169 55%, #0A4FB8 100%);
        padding: 2rem 2.5rem;
        border-radius: 12px;
        margin-bottom: 2rem;
        box-shadow: 0 6px 24px rgba(1, 33, 105, 0.20);
        position: relative;
        overflow: hidden;
    }
    .app-header::after {
        content: "";
        position: absolute;
        inset: 0;
        background: radial-gradient(circle at 90% 20%, rgba(255,255,255,0.08), transparent 60%);
        pointer-events: none;
    }
    .app-header h1 { color: white; font-size: 2rem; font-weight: 700; margin: 0 0 0.3rem 0; letter-spacing: -0.02em; }
    .app-header p  { color: rgba(255,255,255,0.85); font-size: 1rem; margin: 0; }
    .app-header .brand-tag {
        display: inline-block; margin-top: 0.75rem;
        padding: 0.2rem 0.7rem;
        background: rgba(255,255,255,0.12);
        border: 1px solid rgba(255,255,255,0.25);
        border-radius: 999px; color: #FFFFFF; font-size: 0.75rem;
        font-weight: 600; letter-spacing: 0.08em; text-transform: uppercase;
    }
    .app-header .ai-tag {
        display: inline-block; margin-left: 0.5rem;
        padding: 0.2rem 0.7rem;
        background: linear-gradient(135deg, #06b6d4, #3b82f6);
        border: 1px solid rgba(255,255,255,0.4);
        border-radius: 999px; color: #FFFFFF; font-size: 0.75rem;
        font-weight: 700; letter-spacing: 0.05em;
    }

    .metric-row { display: flex; gap: 1rem; margin: 1.5rem 0; }
    .metric-card {
        background: #FFFFFF; border: 1px solid #E2E6EE; border-radius: 10px;
        padding: 1.25rem 1.5rem; flex: 1; text-align: center;
        transition: border-color 0.2s, box-shadow 0.2s;
    }
    .metric-card:hover { border-color: #012169; box-shadow: 0 4px 12px rgba(1, 33, 105, 0.10); }
    .metric-card .metric-value { font-size: 2rem; font-weight: 700; color: #001A52; line-height: 1.2; }
    .metric-card .metric-label { font-size: 0.8rem; color: #6B7280; text-transform: uppercase; letter-spacing: 0.05em; margin-top: 0.25rem; }
    .metric-card.accent .metric-value { color: #012169; }
    .metric-card.success .metric-value { color: #15803D; }
    .metric-card.warning .metric-value { color: #B45309; }
    .metric-card.error   .metric-value { color: #B91C1C; }

    .status-log {
        background: #F4F6FA; border: 1px solid #E2E6EE; border-radius: 8px;
        padding: 1rem; font-family: 'Consolas', 'Monaco', monospace;
        font-size: 0.8rem; max-height: 300px; overflow-y: auto;
        color: #3F4754; line-height: 1.6;
    }
    .status-log .log-info { color: #012169; }
    .status-log .log-success { color: #15803D; }
    .status-log .log-warning { color: #B45309; }
    .status-log .log-error { color: #B91C1C; }
    .status-log .log-step { color: #0A4FB8; font-weight: 600; }

    section[data-testid="stSidebar"] { background: #F4F6FA; border-right: 1px solid #E2E6EE; }
    section[data-testid="stSidebar"] h3 { color: #012169; }

    .stButton > button {
        background: linear-gradient(135deg, #012169, #0A4FB8);
        color: white; border: none; padding: 0.6rem 2rem; font-weight: 600;
        border-radius: 8px; transition: all 0.2s; width: 100%;
    }
    .stButton > button:hover { box-shadow: 0 4px 16px rgba(1, 33, 105, 0.35); transform: translateY(-1px); }

    .stDownloadButton > button {
        background: #FFFFFF; border: 1px solid #E2E6EE; color: #001A52;
        border-radius: 8px; transition: all 0.2s;
    }
    .stDownloadButton > button:hover { border-color: #012169; background: #EFF3FB; }

    .stDataFrame { border-radius: 10px; overflow: hidden; border: 1px solid #E2E6EE; }
    .stTabs [data-baseweb="tab-list"] { gap: 0.5rem; }
    .stTabs [data-baseweb="tab"] { border-radius: 8px 8px 0 0; padding: 0.5rem 1.5rem; }
    .stTabs [aria-selected="true"] { color: #012169 !important; }

    .stTextInput > div > div > input { background: #FFFFFF; border-color: #E2E6EE; color: #001A52; border-radius: 8px; }
    .stTextInput > div > div > input:focus { border-color: #012169; box-shadow: 0 0 0 1px #012169; }

    .streamlit-expanderHeader { background: #F4F6FA; border-radius: 8px; }
    .main .stMarkdown h3 { color: #001A52; font-weight: 600; border-bottom: 2px solid #E2E6EE; padding-bottom: 0.5rem; margin-top: 1.5rem; }

    .footer-brand {
        margin-top: 3rem; padding-top: 1.25rem;
        border-top: 1px solid #E2E6EE; text-align: center;
        color: #898D8D; font-size: 0.78rem; letter-spacing: 0.05em;
    }
    .footer-brand strong { color: #012169; font-weight: 700; letter-spacing: 0.08em; }

    #MainMenu { visibility: hidden; }
    footer { visibility: hidden; }
    header { visibility: hidden; }
    header [data-testid="stExpandSidebarButton"],
    button[data-testid="stExpandSidebarButton"] {
        visibility: visible !important;
        z-index: 999999;
    }
</style>
""", unsafe_allow_html=True)


# ── Session state ─────────────────────────────────────────────────────

_loaded = load_settings()

for key, default in [
    ("results", None),
    ("log_lines", []),
    ("is_running", False),
    ("run_stats", {}),
    # Env var takes precedence over saved file (lets a power user override
    # via env if they want). Otherwise use the persisted value.
    ("ai_key", os.getenv("GEMINI_API_KEY") or _loaded["gemini_api_key"]),
    ("output_format", _loaded["output_format"]),
    ("gemini_model", _loaded["gemini_model"]),
]:
    if key not in st.session_state:
        st.session_state[key] = default


# ── Logging bridge ────────────────────────────────────────────────────

class StreamlitLogHandler(logging.Handler):
    def emit(self, record):
        msg = html_mod.escape(self.format(record))
        css = {"INFO": "log-info", "WARNING": "log-warning", "ERROR": "log-error"}.get(record.levelname, "")
        st.session_state.log_lines.append(f'<span class="{css}">{msg}</span>')
        if len(st.session_state.log_lines) > 200:
            st.session_state.log_lines = st.session_state.log_lines[-200:]


def setup_ui_logging():
    fmt = logging.Formatter("[%(asctime)s] %(message)s", datefmt="%H:%M:%S")
    handler = StreamlitLogHandler()
    handler.setFormatter(fmt)
    handler.setLevel(logging.INFO)
    root = logging.getLogger()
    root.setLevel(logging.INFO)
    for h in root.handlers[:]:
        root.removeHandler(h)
    root.addHandler(handler)
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    log_file = LOG_DIR / f"scraper_ai_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    fh = logging.FileHandler(log_file, encoding="utf-8")
    fh.setFormatter(logging.Formatter(
        "[%(asctime)s] [%(levelname)s] [%(name)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    ))
    fh.setLevel(logging.DEBUG)
    root.addHandler(fh)


def build_planner() -> Planner | None:
    key = (st.session_state.get("ai_key") or "").strip()
    if not key:
        return None
    model = (st.session_state.get("gemini_model") or "gemini-2.5-flash").strip()
    return Planner(PlannerConfig(api_key=key, model=model))


# ── Runner ────────────────────────────────────────────────────────────

def run_scraper(companies, progress_bar, status_text, counter_text=None,
                manual_url=None, output_fmt="both"):
    setup_ui_logging()
    logger = logging.getLogger("main")
    st.session_state.log_lines = []

    planner = build_planner()
    if planner is None:
        logger.info("No API key — running standard pipeline")
    else:
        logger.info("AI planner enabled")

    all_locations, errors = [], []
    companies_done = 0
    start_time = time.time()
    total = len(companies)

    def _update_counter():
        if counter_text is None:
            return
        counter_text.markdown(
            f"**{len(all_locations):,}** locations found &nbsp;·&nbsp; "
            f"**{companies_done}** / **{total}** companies done"
        )

    def progress_callback(step, company, detail=""):
        labels = {
            "discovery": "Discovering store locator URL",
            "rendering": "Rendering page",
            "extracting": "Extracting locations",
            "cleaning": "Cleaning & normalizing",
            "validating": "Validating results",
        }
        label = labels.get(step, step)
        if detail:
            status_text.markdown(f"**{company}** — {label} · `{detail}`")
        else:
            status_text.markdown(f"**{company}** — {label}")

    def _save_progress():
        elapsed = time.time() - start_time
        st.session_state.run_stats = {
            "total_companies": len(companies),
            "successful": companies_done - len(errors),
            "total_locations": len(all_locations),
            "errors": list(errors),
            "elapsed": elapsed,
        }
        if all_locations:
            df = pd.DataFrame(all_locations)
            for col in COLUMNS:
                if col not in df.columns:
                    df[col] = ""
            st.session_state.results = df[COLUMNS]

    for i, company in enumerate(companies):
        pct = int(i / total * 100)
        progress_bar.progress(i / total, text=f"{pct}% — Processing {i+1}/{total}: {company}")
        try:
            locs = scrape_company_ai(company, planner=planner, manual_url=manual_url,
                                     progress_callback=progress_callback)
            all_locations.extend(locs)
            if not locs:
                errors.append((company, "No locations found"))
        except Exception as e:
            logger.error("[%s] Error: %s", company, e)
            errors.append((company, str(e)))

        companies_done += 1
        _update_counter()
        _save_progress()

        if i < total - 1:
            delay = random.uniform(LONG_DELAY_MIN, LONG_DELAY_MAX) if (i + 1) % LONG_DELAY_EVERY_N == 0 \
                else random.uniform(REQUEST_DELAY_MIN, REQUEST_DELAY_MAX)
            time.sleep(delay)

    progress_bar.progress(1.0, text="100% — Complete")

    _save_progress()
    if all_locations:
        base_name = f"locations_ai_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        export_results(all_locations, output_fmt, base_name)
    else:
        st.session_state.results = pd.DataFrame()


# ── Header ────────────────────────────────────────────────────────────

st.markdown("""
<div class="app-header">
    <span class="brand-tag">By INDATEL Labs</span>
    <span class="ai-tag">AI&nbsp;POWERED</span>
    <h1 style="margin-top: 0.6rem;">Location Scraper</h1>
    <p>Extract physical locations for any retail chain — AI handles discovery and extraction</p>
</div>
""", unsafe_allow_html=True)


# ── Sidebar ───────────────────────────────────────────────────────────

with st.sidebar:
    st.markdown("### Settings")

    with st.expander("⚙️  AI & Output", expanded=True):
        key_input = st.text_input(
            "AI API Key",
            type="password",
            value=st.session_state.ai_key,
            placeholder="Paste your INDATEL Labs key",
            help=(
                "Contact INDATEL Labs to receive an API key. "
                "The key is stored locally on this machine only. "
                "Without a key, the app falls back to the standard scraper."
            ),
            key="ai_key_input",
        )

        format_options = ["Both (CSV + Excel)", "CSV Only", "Excel Only"]
        try:
            fmt_idx = format_options.index(st.session_state.output_format)
        except ValueError:
            fmt_idx = 0
        format_input = st.selectbox(
            "Output Format",
            format_options,
            index=fmt_idx,
            key="output_format_input",
        )

        # Auto-persist when anything changed
        if (
            key_input != st.session_state.ai_key
            or format_input != st.session_state.output_format
        ):
            st.session_state.ai_key = key_input
            st.session_state.output_format = format_input
            save_settings({
                "gemini_api_key": key_input,
                "gemini_model": st.session_state.gemini_model,
                "output_format": format_input,
            })

        if st.session_state.ai_key:
            st.success("AI enabled")
        else:
            st.info("Don't have a key? **Contact INDATEL Labs** to get one.")

    output_format = st.session_state.output_format  # used downstream

    st.markdown("---")
    st.markdown("### Tips")
    st.markdown("""
    - Use the **exact company name** as it appears on their website
    - Works best with US-based retail chains
    - With AI enabled, new/unknown chains usually work first try
    - Large chains (1000+ locations) may take 1-2 minutes
    """)

    st.markdown("---")
    _update_available, _latest_v, _release_url = check_for_update()
    if _update_available and _release_url:
        st.warning(
            f"**Update available:** v{_latest_v}\n\n"
            f"[Download new installer]({_release_url})"
        )
    st.caption(f"v{__version__}  •  INDATEL Labs")


# ── Main content ──────────────────────────────────────────────────────

tab_single, tab_batch = st.tabs(["Single Company", "Batch (Excel / CSV Upload)"])

with tab_single:
    with st.form(key="single_form", clear_on_submit=False, border=False):
        col_input, col_btn = st.columns([3, 1])
        with col_input:
            company_name = st.text_input(
                "Company Name",
                placeholder="e.g. Aldi, Starbucks, Dollar Tree, Sprouts, Joe's Pizza...",
                label_visibility="collapsed",
                key="single_company_input",
            )
        with col_btn:
            st.markdown("<div style='height: 0.45rem'></div>", unsafe_allow_html=True)
            run_single = st.form_submit_button("Scrape Locations", use_container_width=True)
        manual_url = st.text_input(
            "Locator URL (optional)",
            placeholder="https://locations.example.com/  — paste if you already know the URL",
            key="single_manual_url",
            help="If you already know the chain's store locator URL, paste it here to skip discovery.",
        )

with tab_batch:
    uploaded_file = st.file_uploader(
        "Upload an Excel or CSV file with company names in the first column",
        type=["xlsx", "csv"], key="batch_upload",
    )
    companies_from_file = []
    if uploaded_file:
        _HEADER_WORDS = {"company", "company_name", "name", "company name"}
        fname = (uploaded_file.name or "").lower()
        try:
            if fname.endswith(".csv"):
                import csv as csv_mod
                raw = uploaded_file.read().decode("utf-8-sig")
                reader = csv_mod.reader(io.StringIO(raw))
                for row in reader:
                    if row:
                        v = (row[0] or "").strip()
                        if v and v.lower() not in _HEADER_WORDS:
                            companies_from_file.append(v)
            else:
                from openpyxl import load_workbook
                wb = load_workbook(io.BytesIO(uploaded_file.read()), read_only=True)
                ws = wb.active
                for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
                    v = row[0]
                    if v and isinstance(v, str):
                        vc = v.strip()
                        if vc and vc.lower() not in _HEADER_WORDS:
                            companies_from_file.append(vc)
                wb.close()
        except Exception as e:
            st.error(f"Could not read the uploaded file: {e}")
            companies_from_file = []
        if companies_from_file:
            st.success(f"Found **{len(companies_from_file)}** companies: {', '.join(companies_from_file[:10])}")
            if len(companies_from_file) > 10:
                st.caption(f"...and {len(companies_from_file) - 10} more")
        else:
            st.error("No company names found in the first column.")
    run_batch = st.button("Scrape All Companies", key="run_batch", use_container_width=True)


# ── Execute scraping ──────────────────────────────────────────────────

def _company_from_url(u: str) -> str:
    from urllib.parse import urlparse
    host = (urlparse(u).hostname or "").lower()
    if host.startswith("www."):
        host = host[4:]
    if host.startswith(("locations.", "stores.", "find.", "store.", "restaurants.")):
        host = host.split(".", 1)[1]
    parts = host.split(".")
    brand = parts[-2] if len(parts) >= 2 else (parts[0] if parts else "company")
    return brand.replace("-", " ").title() if brand else "Company"


# Reset flag if a previous run was cancelled mid-flight
if st.session_state.is_running:
    st.session_state.is_running = False

companies_to_scrape: list[str] = []
single_manual_url: str | None = None

if run_single:
    url_in = (manual_url or "").strip()
    name_in = (company_name or "").strip()
    if name_in:
        companies_to_scrape = [name_in]
        single_manual_url = url_in or None
    elif url_in:
        companies_to_scrape = [_company_from_url(url_in)]
        single_manual_url = url_in
elif run_batch and uploaded_file and companies_from_file:
    companies_to_scrape = companies_from_file

if companies_to_scrape:
    st.session_state.is_running = True
    st.session_state.results = None
    st.session_state.run_stats = {}

    st.markdown("---")
    st.markdown("### Scraping Progress")
    progress_bar = st.progress(0, text="Starting...")
    status_text = st.empty()
    counter_text = st.empty()

    if len(companies_to_scrape) > 1:
        st.caption("Click **Cancel** to stop after the current company finishes.")
        st.button("Cancel Scraping", key="cancel_scrape", type="secondary")

    fmt_map = {"Both (CSV + Excel)": "both", "CSV Only": "csv", "Excel Only": "excel"}
    chosen_fmt = fmt_map.get(output_format, "both")

    with st.spinner(""):
        run_scraper(companies_to_scrape, progress_bar, status_text,
                    counter_text=counter_text,
                    manual_url=single_manual_url, output_fmt=chosen_fmt)

    st.session_state.is_running = False
    status_text.empty()
    counter_text.empty()


# ── Results display ───────────────────────────────────────────────────

if st.session_state.run_stats:
    stats = st.session_state.run_stats
    st.markdown("---")
    err_count = len(stats.get("errors", []))
    st.markdown(f"""
    <div class="metric-row">
        <div class="metric-card accent">
            <div class="metric-value">{stats.get('successful', 0)}/{stats.get('total_companies', 0)}</div>
            <div class="metric-label">Companies Processed</div>
        </div>
        <div class="metric-card success">
            <div class="metric-value">{stats.get('total_locations', 0):,}</div>
            <div class="metric-label">Locations Found</div>
        </div>
        <div class="metric-card {'warning' if err_count else 'success'}">
            <div class="metric-value">{err_count}</div>
            <div class="metric-label">Errors</div>
        </div>
        <div class="metric-card">
            <div class="metric-value">{stats.get('elapsed', 0):.0f}s</div>
            <div class="metric-label">Time Elapsed</div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    if stats.get("errors"):
        with st.expander(f"Errors ({len(stats['errors'])})"):
            for comp, err in stats["errors"]:
                st.markdown(f"- **{comp}**: {err}")


if st.session_state.results is not None and not st.session_state.results.empty:
    df = st.session_state.results

    st.markdown("### Results")
    col_f1, col_f2, col_f3, col_f4 = st.columns(4)
    with col_f1:
        companies_in = sorted(df["company_name"].unique())
        selected_company = st.selectbox("Company", ["All"] + companies_in, key="filter_company")
    with col_f2:
        states_in = sorted([s for s in df["state"].dropna().unique() if s])
        selected_state = st.selectbox("State", ["All"] + states_in, key="filter_state")
    with col_f3:
        quality_opts = sorted([q for q in df["data_quality"].dropna().unique() if q])
        selected_quality = st.selectbox("Data Quality", ["All"] + list(quality_opts), key="filter_quality")
    with col_f4:
        search_text = st.text_input("Search", placeholder="Filter by address, city...", key="filter_search")

    filtered = df.copy()
    if selected_company != "All":
        filtered = filtered[filtered["company_name"] == selected_company]
    if selected_state != "All":
        filtered = filtered[filtered["state"] == selected_state]
    if selected_quality != "All":
        filtered = filtered[filtered["data_quality"] == selected_quality]
    if search_text:
        mask = filtered.astype(str).apply(
            lambda row: row.str.contains(search_text, case=False, na=False).any(), axis=1)
        filtered = filtered[mask]

    st.caption(f"Showing {len(filtered):,} of {len(df):,} locations")

    display_cols = ["company_name", "location_name", "street_address", "city", "state",
                    "zip_code", "phone_number", "data_quality"]
    st.dataframe(
        filtered[display_cols],
        use_container_width=True, height=450,
        column_config={
            "company_name":   st.column_config.TextColumn("Company", width="medium"),
            "location_name":  st.column_config.TextColumn("Location", width="medium"),
            "street_address": st.column_config.TextColumn("Address", width="large"),
            "city":           st.column_config.TextColumn("City", width="small"),
            "state":          st.column_config.TextColumn("State", width="small"),
            "zip_code":       st.column_config.TextColumn("ZIP", width="small"),
            "phone_number":   st.column_config.TextColumn("Phone", width="medium"),
            "data_quality":   st.column_config.TextColumn("Quality", width="small"),
        },
    )

    if len(filtered) > 10:
        with st.expander("Location Distribution by State"):
            st.bar_chart(filtered["state"].value_counts().head(20))

    st.markdown("### Download Results")
    col_d1, col_d2, col_d3 = st.columns(3)
    with col_d1:
        st.download_button("Download CSV", data=filtered.to_csv(index=False).encode("utf-8"),
                           file_name=f"locations_ai_{datetime.now().strftime('%Y%m%d')}.csv",
                           mime="text/csv", use_container_width=True)
    with col_d2:
        buf = io.BytesIO(); filtered.to_excel(buf, index=False, engine="openpyxl"); buf.seek(0)
        st.download_button("Download Excel", data=buf.getvalue(),
                           file_name=f"locations_ai_{datetime.now().strftime('%Y%m%d')}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)
    with col_d3:
        st.download_button("Download All (Unfiltered)", data=df.to_csv(index=False).encode("utf-8"),
                           file_name=f"all_locations_ai_{datetime.now().strftime('%Y%m%d')}.csv",
                           mime="text/csv", use_container_width=True)

    if st.session_state.log_lines:
        with st.expander("Activity Log"):
            log_html = "<br>".join(st.session_state.log_lines[-100:])
            st.markdown(f'<div class="status-log">{log_html}</div>', unsafe_allow_html=True)

elif st.session_state.results is not None and st.session_state.results.empty:
    st.info("No locations were found. Try a different company name or check the activity log.")
    if st.session_state.log_lines:
        with st.expander("Activity Log"):
            log_html = "<br>".join(st.session_state.log_lines[-100:])
            st.markdown(f'<div class="status-log">{log_html}</div>', unsafe_allow_html=True)


# ── Footer ────────────────────────────────────────────────────────────

st.markdown("""
<div class="footer-brand">
    By <strong>INDATEL LABS</strong> · AI-Powered Edition
</div>
""", unsafe_allow_html=True)
