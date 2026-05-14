import csv
import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config.settings import OUTPUT_DIR

logger = logging.getLogger("scraper.exporter")

COLUMNS = [
    "company_name",
    "location_name",
    "street_address",
    "city",
    "state",
    "zip_code",
    "country",
    "full_address",
    "phone_number",
    "hours_of_operation",
    "location_type",
    "latitude",
    "longitude",
    "location_url",
    "source_url",
    "scraped_at",
    "data_quality",
]


def _ensure_output_dir():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def export_csv(locations: list[dict], filename: str | None = None) -> Path:
    _ensure_output_dir()
    if not filename:
        filename = f"locations_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    path = OUTPUT_DIR / filename

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=COLUMNS, extrasaction="ignore")
        writer.writeheader()
        for loc in locations:
            writer.writerow({col: loc.get(col, "") for col in COLUMNS})

    logger.info("CSV exported: %s (%d rows)", path, len(locations))
    return path


def export_excel(locations: list[dict], filename: str | None = None) -> Path:
    _ensure_output_dir()
    if not filename:
        filename = f"locations_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    path = OUTPUT_DIR / filename

    wb = Workbook()
    ws = wb.active
    ws.title = "Locations"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write headers
    display_headers = [col.replace("_", " ").title() for col in COLUMNS]
    for col_idx, header in enumerate(display_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Quality color coding
    quality_fills = {
        "complete": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "partial": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "address_only": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    }

    # Write data
    for row_idx, loc in enumerate(locations, 2):
        for col_idx, col_name in enumerate(COLUMNS, 1):
            value = loc.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=(col_name == "hours_of_operation"))

            # Color the data_quality column
            if col_name == "data_quality" and value in quality_fills:
                cell.fill = quality_fills[value]

    # Auto-width columns
    for col_idx in range(1, len(COLUMNS) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(display_headers[col_idx - 1])
        for row in ws.iter_rows(min_row=2, max_row=min(len(locations) + 1, 50), min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), 50))
        ws.column_dimensions[col_letter].width = max_len + 3

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(locations) + 1}"

    wb.save(path)
    logger.info("Excel exported: %s (%d rows)", path, len(locations))
    return path


def export_results(
    locations: list[dict],
    fmt: str = "both",
    base_filename: str | None = None,
) -> dict[str, Path]:
    """
    Export locations to CSV and/or Excel.
    fmt: "csv", "excel", or "both"
    Returns dict of format -> file path.
    """
    if not base_filename:
        base_filename = f"locations_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

    paths = {}
    if fmt in ("csv", "both"):
        paths["csv"] = export_csv(locations, f"{base_filename}.csv")
    if fmt in ("excel", "both"):
        paths["excel"] = export_excel(locations, f"{base_filename}.xlsx")
    return paths
