"""
Retail Chain Location Scraper
Extracts physical locations for retail chains from their official websites.

Usage:
    python main.py --company "Starbucks"
    python main.py --file companies.xlsx
    python main.py --company "McDonald's" --output ./results --format both
"""

import argparse
import logging
import sys
import time
import random
from datetime import datetime
from pathlib import Path

from config.settings import (
    OUTPUT_DIR,
    LOG_DIR,
    LOG_LEVEL,
    MAX_LOCATIONS,
    REQUEST_DELAY_MIN,
    REQUEST_DELAY_MAX,
    LONG_DELAY_MIN,
    LONG_DELAY_MAX,
    LONG_DELAY_EVERY_N,
)
from scraper.discovery import discover_locator_url
from scraper.renderer import render_page
from scraper.extractor import extract_locations
from scraper.cleaner import clean_locations
from scraper.exporter import export_results


def setup_logging():
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    log_file = LOG_DIR / f"scraper_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"

    fmt = logging.Formatter(
        "[%(asctime)s] [%(levelname)s] [%(name)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    file_handler = logging.FileHandler(log_file, encoding="utf-8")
    file_handler.setFormatter(fmt)
    file_handler.setLevel(logging.DEBUG)

    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setFormatter(fmt)
    console_handler.setLevel(getattr(logging, LOG_LEVEL, logging.INFO))

    root = logging.getLogger()
    root.setLevel(logging.DEBUG)
    root.addHandler(file_handler)
    root.addHandler(console_handler)

    return log_file


def load_companies_from_excel(filepath: str) -> list[str]:
    from openpyxl import load_workbook

    path = Path(filepath)
    if not path.exists():
        print(f"Error: File not found: {filepath}")
        sys.exit(1)

    wb = load_workbook(path, read_only=True)
    ws = wb.active

    companies = []
    for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
        val = row[0]
        if val and isinstance(val, str):
            cleaned = val.strip()
            if cleaned and cleaned.lower() not in ("company", "company_name", "name", "company name"):
                companies.append(cleaned)

    wb.close()

    if not companies:
        print("Error: No company names found in the first column of the Excel file.")
        sys.exit(1)

    return companies


def scrape_company(company_name: str, logger: logging.Logger) -> list[dict]:
    logger.info("=" * 60)
    logger.info("Processing: %s", company_name)
    logger.info("=" * 60)

    # Step 1: Discover store locator URL
    logger.info("[%s] Step 1: Discovering store locator URL...", company_name)
    url, method = discover_locator_url(company_name)
    if not url:
        logger.warning("[%s] Could not find store locator URL. Skipping.", company_name)
        return []
    logger.info("[%s] Found URL via %s: %s", company_name, method, url)

    time.sleep(random.uniform(REQUEST_DELAY_MIN, REQUEST_DELAY_MAX))

    # Step 2: Render the page
    logger.info("[%s] Step 2: Rendering page...", company_name)
    result = render_page(url)
    if not result:
        logger.warning("[%s] Failed to render page. Skipping.", company_name)
        return []
    logger.info("[%s] Page rendered via %s (%d bytes)", company_name, result.method, len(result.html))

    # Step 3: Extract locations
    logger.info("[%s] Step 3: Extracting locations...", company_name)
    raw_locations = extract_locations(result, company_name)

    # If static rendering yielded 0 results, retry with Playwright (API sniffing)
    if not raw_locations and result.method == "static":
        logger.info("[%s] Static extraction found nothing — retrying with Playwright...", company_name)
        time.sleep(random.uniform(REQUEST_DELAY_MIN, REQUEST_DELAY_MAX))
        pw_result = render_page(url, force_playwright=True)
        if pw_result:
            raw_locations = extract_locations(pw_result, company_name)

    # Try the locations.{brand}.com subdomain directly — many chains host a
    # Yext/Uberall directory there even when their main /store-locator page
    # is a search-only interface.
    if not raw_locations:
        from scraper.discovery import _slug_candidates
        from scraper.extractor import extract_from_directory_tree
        import requests
        for slug in _slug_candidates(company_name):
            alt_url = f"https://locations.{slug}.com/index.html"
            try:
                logger.info("[%s] Trying directory subdomain: %s", company_name, alt_url)
                r = requests.get(alt_url, headers={"User-Agent": "Mozilla/5.0"}, timeout=12, allow_redirects=True)
                if r.status_code != 200 or len(r.text) < 1000:
                    continue
                dir_locs = extract_from_directory_tree(r.text, alt_url)
                if dir_locs:
                    raw_locations = dir_locs
                    logger.info("[%s] Directory subdomain extraction: %d locations", company_name, len(dir_locs))
                    break
            except requests.RequestException:
                continue

    # Last resort: try direct API probing on common endpoint patterns
    if not raw_locations:
        from scraper.extractor import probe_direct_apis
        logger.info("[%s] Trying direct API endpoint probing...", company_name)
        raw_locations = probe_direct_apis(url, company_name)

    if not raw_locations:
        logger.warning("[%s] No locations extracted.", company_name)
        return []
    logger.info("[%s] Extracted %d raw locations", company_name, len(raw_locations))

    # Step 4: Clean and normalize
    logger.info("[%s] Step 4: Cleaning and normalizing...", company_name)
    cleaned = clean_locations(raw_locations, company_name)
    logger.info("[%s] Final count: %d locations", company_name, len(cleaned))

    if MAX_LOCATIONS > 0 and len(cleaned) > MAX_LOCATIONS:
        logger.info("[%s] Trimming to max %d locations", company_name, MAX_LOCATIONS)
        cleaned = cleaned[:MAX_LOCATIONS]

    return cleaned


def main():
    parser = argparse.ArgumentParser(
        description="Retail Chain Location Scraper",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python main.py --company "Starbucks"
  python main.py --file companies.xlsx
  python main.py --company "McDonald's" --output ./results --format both
  python main.py --company "Chase Bank" --format excel
        """,
    )
    parser.add_argument(
        "--company", "-c",
        type=str,
        help="Single company name to scrape",
    )
    parser.add_argument(
        "--file", "-f",
        type=str,
        help="Path to Excel file with company names (first column)",
    )
    parser.add_argument(
        "--output", "-o",
        type=str,
        help=f"Output directory (default: {OUTPUT_DIR})",
    )
    parser.add_argument(
        "--format",
        choices=["csv", "excel", "both"],
        default="both",
        help="Output format (default: both)",
    )

    args = parser.parse_args()

    if not args.company and not args.file:
        parser.print_help()
        print("\nError: Provide either --company or --file")
        sys.exit(1)

    if args.output:
        import config.settings as settings
        settings.OUTPUT_DIR = Path(args.output)

    log_file = setup_logging()
    logger = logging.getLogger("main")
    logger.info("Location Scraper started")
    logger.info("Log file: %s", log_file)

    # Build company list
    if args.file:
        companies = load_companies_from_excel(args.file)
        logger.info("Loaded %d companies from %s", len(companies), args.file)
    else:
        companies = [args.company]

    # Process each company
    all_locations = []
    errors = []
    start_time = time.time()

    for i, company in enumerate(companies):
        try:
            locations = scrape_company(company, logger)
            all_locations.extend(locations)
            if not locations:
                errors.append((company, "No locations found"))
        except KeyboardInterrupt:
            logger.info("Interrupted by user. Saving collected data...")
            break
        except Exception as e:
            logger.error("[%s] Unhandled error: %s", company, e, exc_info=True)
            errors.append((company, str(e)))

        # Rate limiting between companies
        if i < len(companies) - 1:
            if (i + 1) % LONG_DELAY_EVERY_N == 0:
                delay = random.uniform(LONG_DELAY_MIN, LONG_DELAY_MAX)
                logger.info("Long delay: %.1fs (every %d companies)", delay, LONG_DELAY_EVERY_N)
            else:
                delay = random.uniform(REQUEST_DELAY_MIN, REQUEST_DELAY_MAX)
            time.sleep(delay)

    elapsed = time.time() - start_time

    # Export results
    if all_locations:
        base_name = f"locations_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        paths = export_results(all_locations, args.format, base_name)
        output_str = ", ".join(str(p) for p in paths.values())
    else:
        output_str = "(no data to export)"

    # Summary
    successful = len(companies) - len(errors)
    print()
    print("=" * 50)
    print(f"  Companies processed: {successful}/{len(companies)}")
    print(f"  Total locations found: {len(all_locations):,}")
    if errors:
        print(f"  Companies with errors: {len(errors)} (see log)")
        for comp, err in errors:
            print(f"    - {comp}: {err}")
    print(f"  Output saved to: {output_str}")
    print(f"  Time elapsed: {elapsed:.1f}s")
    print(f"  Log file: {log_file}")
    print("=" * 50)


if __name__ == "__main__":
    main()
